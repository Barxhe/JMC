import os
import re
import openpyxl
import sqlite3 as sql
from win32com.client import Dispatch
import time
import gc

filelist =[]
db_path = 'D:\dbtool'
os.chdir(db_path)
conn = sql.connect('JMC.db')
cursor = conn.cursor()
sql = "insert into JMC values(?,?,?,?,?,?,?,?,?)"

#############table_creat####################

#########################################File_list##########################################
# class XlsClass:
#     def __init__(self, xlApp, filename=None ,*,Visible=False ,Alerts=False):
#         self.xlApp = xlApp
#         self.xlApp.Visible = Visible


#     def otherMethod(self):
#         # ....

#     def close(self):
#         self.xlBook.Close()
#         self.xlBook = None
#         self.xlApp = None
#         # don't do anything with self.xlApp or self

# xlApp = win32com.client.Dispatch('Excel.Application')
# for fname in filelist:
#     xlbook = XlsClass(xlApp, fname)
#     # do something with xlbook
#     xlbook.close()

######################################DB_FORMAT#####################################
class Format_1():
    def __init__(self,i,j):
        self.i = i
        self.j = j
    
    def row_format(self):
        content1 =  (work_sheet.cell(row = 2 ,column = 4).value,work_sheet.cell(row = self.j ,column = 2).value,
            work_sheet.cell(row = self.i+self.j,column = 4).value,work_sheet.cell(row = self.i+self.j,column = 5).value,                 
            work_sheet.cell(row = self.i+self.j,column = 6).value,work_sheet.cell(row = self.i+self.j,column = 7).value,
            work_sheet.cell(row = self.i+self.j,column = 8).value,'Null',work_sheet.cell(row = 4 ,column = 10).value
            )
        return content1

class Format_2():
    def __init__(self,i,j):
        self.i = i
        self.j = j
    
    def row_format(self):
        content2 =  (work_sheet.cell(row = 2 ,column = 4).value,work_sheet.cell(row = 25 ,column = 2).value,
            work_sheet.cell(self.j,column = 4).value,work_sheet.cell(row = self.i+self.j,column = 5).value,                 
            work_sheet.cell(row = self.i+self.j,column = 6).value,work_sheet.cell(row = self.i+self.j,column = 7).value,
            work_sheet.cell(row = self.i+self.j,column = 8).value,'Null',work_sheet.cell(row = 4 ,column = 10).value
            )
        return content2

class Format_3():
    def __init__(self,i,j):
        self.i = i
        self.j = j
    
    def row_format(self):
        for rows in range(0,8):
            if rows == 0:
                content3_1 =  (work_sheet.cell(row = 2 ,column = 4).value,work_sheet.cell(row = self.j ,column = 2).value,
                    work_sheet.cell(self.j,column = 4).value,work_sheet.cell(row = rows+self.j,column = 5).value,                 
                    work_sheet.cell(row = rows+self.j,column = 6).value,work_sheet.cell(row = rows+self.j,column = 7).value,
                    work_sheet.cell(row = self.i+self.j,column = 8).value,'Null',work_sheet.cell(row = 4 ,column = 10).value
                    )
                cursor.execute(sql,(content3_1))
            elif rows <7:
                content3_2 =  (work_sheet.cell(row = 2 ,column = 4).value,work_sheet.cell(row = self.j,column = 2).value,
                    work_sheet.cell(row = rows+self.j,column = 4).value,work_sheet.cell(row = rows+self.j,column = 5).value,                 
                    work_sheet.cell(row = rows+self.j,column = 6).value,work_sheet.cell(row = rows+self.j,column = 7).value,
                    work_sheet.cell(row = self.i+self.j,column = 8).value,'Null',work_sheet.cell(row = 4 ,column = 10).value
                    # work_sheet.cell(row = 6 ,column = 8).value,work_sheet.cell(row = 6 ,column = 9).value,
                )
                cursor.execute(sql,(content3_2))            
            else:
                content3_3 =  (work_sheet.cell(row = 2 ,column = 4).value,work_sheet.cell(row = self.j ,column = 2).value,
                    work_sheet.cell(rows+self.j,column = 4).value,work_sheet.cell(row = rows+self.j,column = 5).value,                 
                    work_sheet.cell(row = rows+self.j,column = 6).value,work_sheet.cell(row = rows+self.j,column = 7).value,
                    work_sheet.cell(row = self.i+self.j,column = 8).value,'Null',work_sheet.cell(row = 4 ,column = 10).value
                    # work_sheet.cell(row = 6 ,column = 8).value,work_sheet.cell(row = 6 ,column = 9).value,
                    )
                cursor.execute(sql,(content3_3)) 

#####################################Sys_init_1###################################
def Sy():
    Sys_init_1 = Format_1(0,6)
    cursor.execute(sql,(Sys_init_1.row_format()))

####################################Tcu_ect##########################
def Te():   
    for n in range(0,8):    ###columns_delta
        Tcu_ect_cycle = Format_1(n,7)
        cursor.execute(sql,(Tcu_ect_cycle.row_format()))
    
##############################KitTankFilled#######################
def Of():
    Oil_Filed = Format_1(0,15)
    cursor.execute(sql,(Oil_Filed.row_format()))

##############################Purge_1############################3#
def P1():
    for n in range(0,2):
        Purge_1 = Format_1(n,16)
        cursor.execute(sql,(Purge_1.row_format()))

##############################Ep_acc_check#############################
def Eac():
    for n in range(0,5):    ###columns_delta
        Ep_acc_check= Format_1(n,18)
        cursor.execute(sql,(Ep_acc_check.row_format()))
      
##############################Purge_2############################3#
def P2():
    for n in range(0,2):
        Purge_2 = Format_1(n,23)
        cursor.execute(sql,(Purge_2.row_format()))

#####################################Grid_self_tun###################################
def Gst():
    Grid_IDD = Format_1(0,25)
    # Grid_Cst = Format_1(0,26)
    cursor.execute(sql,(Grid_IDD.row_format()))
    # cursor.execute(sql,(Grid_Cst.row_format()))

    for n in range(0,8):    ###columns_delta
        Grid_Gcc = Format_2(n,27)

        cursor.execute(sql,(Grid_Gcc.row_format()))
  
    for n in range(0,8):    ###columns_delta
        Grid_Scc = Format_2(n,35)
        cursor.execute(sql,(Grid_Scc.row_format()))  
        
    for n in range(18,23):
        Grid_EV_id = Format_1(n,25)
        cursor.execute(sql,(Grid_EV_id.row_format()))

#####################################Leak_test###################################
def Leak():
    for n in range(0,8):    ###columns_delta
        Leak_test = Format_1(n,48)
        cursor.execute(sql,(Leak_test.row_format()))

#####################################Clutch_test###################################
def Ct():
    Cluth_test = Format_3(0,56)
    Cluth_test.row_format()

#####################################Shift_test###################################
def St():
    Shift_test = Format_3(0,64)
    Shift_test.row_format()

#####################################Gear_test###################################
def Gt():
    Gear_test = Format_3(0,72)
    Gear_test.row_format()


#####################################chdir###################################
for file in os.listdir():
    if re.match('REPORT_',file):
        filelist.append(file)


for file_path in filelist:
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(r'D:\\dbtool\\'+file_path)
    xlBook.Save()
    xlBook.Close()
    wb = openpyxl.load_workbook(file_path,data_only=True)
    work_sheet = wb.get_sheet_by_name('REPORT')
    Sy()
    Te()
    Of()
    P1()
    Eac()
    P2()
    Gst()
    Leak()
    Ct()
    St()
    Gt()
cursor.close()
conn.commit()
conn.close()